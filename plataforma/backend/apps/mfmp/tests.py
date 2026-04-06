"""
Tests del módulo MFMP (M17) — Ley 819/2003.

Cubre modelos, servicios, importación FUT, exports y aislamiento multi-tenant.
"""
from __future__ import annotations

from decimal import Decimal
from io import BytesIO

from django.contrib.auth.models import User
from django.test import TestCase
from rest_framework.test import APIClient

from apps.core.models import Entity, Restructuring, UserEntityAccess
from apps.mfmp.models import (
    MFMP,
    MFMPIncomeProjection,
    MFMPExpenseProjection,
    MFMPDebtProjection,
    MFMPScenario,
    IncomeConcept,
    ExpenseConcept,
)
from apps.mfmp.services import (
    get_projection_matrix,
    calculate_law_617_by_year,
    calculate_law_358_by_year,
    simulate_plan_impact,
    import_fut_excel,
    generate_fiscal_impact_sheet,
)


def _make_entity(name='Alcaldía Test', category='6', order='MUNICIPAL'):
    return Entity.objects.create(
        name=name,
        acronym=name[:8],
        order=order,
        municipality_category=category,
        legal_nature='ALCALDIA',
    )


def _make_mfmp(entity, base_year=2026, horizon=3):
    return MFMP.objects.create(
        entity=entity,
        name=f'MFMP {base_year} test',
        base_year=base_year,
        horizon_years=horizon,
    )


def _seed_incomes(mfmp, tributarios=5_000_000_000, no_tributarios=1_000_000_000,
                  sgp=12_000_000_000):
    years = range(mfmp.base_year, mfmp.base_year + mfmp.horizon_years)
    for yr in years:
        MFMPIncomeProjection.objects.create(
            mfmp=mfmp, year=yr,
            concept=IncomeConcept.TRIBUTARIOS,
            amount=Decimal(str(tributarios)),
        )
        MFMPIncomeProjection.objects.create(
            mfmp=mfmp, year=yr,
            concept=IncomeConcept.NO_TRIBUTARIOS,
            amount=Decimal(str(no_tributarios)),
        )
        MFMPIncomeProjection.objects.create(
            mfmp=mfmp, year=yr,
            concept=IncomeConcept.TRANSFERENCIAS_SGP,
            amount=Decimal(str(sgp)),
        )


def _seed_expenses(mfmp, personal=8_000_000_000, generales=3_500_000_000,
                   transferencias=500_000_000, deuda=800_000_000, inversion=5_000_000_000):
    years = range(mfmp.base_year, mfmp.base_year + mfmp.horizon_years)
    for yr in years:
        MFMPExpenseProjection.objects.create(
            mfmp=mfmp, year=yr,
            concept=ExpenseConcept.FUNCIONAMIENTO_PERSONAL,
            amount=Decimal(str(personal)),
        )
        MFMPExpenseProjection.objects.create(
            mfmp=mfmp, year=yr,
            concept=ExpenseConcept.FUNCIONAMIENTO_GENERALES,
            amount=Decimal(str(generales)),
        )
        MFMPExpenseProjection.objects.create(
            mfmp=mfmp, year=yr,
            concept=ExpenseConcept.FUNCIONAMIENTO_TRANSFERENCIAS,
            amount=Decimal(str(transferencias)),
        )
        MFMPExpenseProjection.objects.create(
            mfmp=mfmp, year=yr,
            concept=ExpenseConcept.SERVICIO_DEUDA,
            amount=Decimal(str(deuda)),
        )
        MFMPExpenseProjection.objects.create(
            mfmp=mfmp, year=yr,
            concept=ExpenseConcept.INVERSION,
            amount=Decimal(str(inversion)),
        )


def _seed_debts(mfmp, saldo=4_000_000_000, servicio=500_000_000):
    years = range(mfmp.base_year, mfmp.base_year + mfmp.horizon_years)
    for yr in years:
        MFMPDebtProjection.objects.create(
            mfmp=mfmp, year=yr,
            outstanding_debt=Decimal(str(saldo)),
            debt_service=Decimal(str(servicio)),
            new_disbursements=Decimal('0'),
        )


# ---------------------------------------------------------------------------
# Tests de servicios de proyección
# ---------------------------------------------------------------------------

class ProjectionMatrixTest(TestCase):
    def setUp(self):
        self.entity = _make_entity()
        self.mfmp = _make_mfmp(self.entity, horizon=3)
        _seed_incomes(self.mfmp)
        _seed_expenses(self.mfmp)
        _seed_debts(self.mfmp)

    def test_create_mfmp_with_incomes_expenses(self):
        """get_projection_matrix devuelve totales correctos."""
        matrix = get_projection_matrix(self.mfmp)
        years = [2026, 2027, 2028]
        self.assertEqual(matrix['years'], years)
        # Total ingresos por año = tributarios + no_tributarios + sgp
        for yr in years:
            expected = 5_000_000_000 + 1_000_000_000 + 12_000_000_000
            actual = float(matrix['totals']['income'][str(yr)])
            self.assertAlmostEqual(actual, expected, delta=1)

    def test_matrix_has_all_keys(self):
        matrix = get_projection_matrix(self.mfmp)
        self.assertIn('years', matrix)
        self.assertIn('income_by_concept', matrix)
        self.assertIn('expense_by_concept', matrix)
        self.assertIn('debt', matrix)
        self.assertIn('totals', matrix)
        self.assertIn('income', matrix['totals'])
        self.assertIn('expense', matrix['totals'])

    def test_expense_total_by_year(self):
        matrix = get_projection_matrix(self.mfmp)
        expected = 8_000_000_000 + 3_500_000_000 + 500_000_000 + 800_000_000 + 5_000_000_000
        actual = float(matrix['totals']['expense']['2026'])
        self.assertAlmostEqual(actual, expected, delta=1)


# ---------------------------------------------------------------------------
# Tests Ley 617
# ---------------------------------------------------------------------------

class Law617Test(TestCase):
    def setUp(self):
        self.entity = _make_entity(category='6')  # límite 80%
        self.mfmp = _make_mfmp(self.entity, horizon=2)

    def test_law_617_compliant_when_expenses_low(self):
        """Funcionamiento/ICLD < 80% → compliant True."""
        _seed_incomes(self.mfmp, tributarios=10_000_000_000, no_tributarios=2_000_000_000)
        # Funcionamiento total = 2B (bien por debajo del 80% de 12B = 9.6B)
        MFMPExpenseProjection.objects.create(
            mfmp=self.mfmp, year=2026,
            concept=ExpenseConcept.FUNCIONAMIENTO_PERSONAL,
            amount=Decimal('1500000000'),
        )
        MFMPExpenseProjection.objects.create(
            mfmp=self.mfmp, year=2026,
            concept=ExpenseConcept.FUNCIONAMIENTO_GENERALES,
            amount=Decimal('500000000'),
        )
        result = calculate_law_617_by_year(self.mfmp)
        self.assertTrue(result[2026]['compliant'])
        self.assertLess(result[2026]['ratio'], 0.80)

    def test_law_617_breaks_when_expenses_high(self):
        """Funcionamiento/ICLD > 80% → compliant False."""
        _seed_incomes(self.mfmp, tributarios=1_000_000_000, no_tributarios=500_000_000)
        # ICLD = 1.5B, funcionamiento = 2B → ratio > 1.0 > 80%
        MFMPExpenseProjection.objects.create(
            mfmp=self.mfmp, year=2026,
            concept=ExpenseConcept.FUNCIONAMIENTO_PERSONAL,
            amount=Decimal('2000000000'),
        )
        result = calculate_law_617_by_year(self.mfmp)
        self.assertFalse(result[2026]['compliant'])
        self.assertGreater(result[2026]['ratio'], 0.80)


# ---------------------------------------------------------------------------
# Tests de simulación de impacto
# ---------------------------------------------------------------------------

class SimulatePlanImpactTest(TestCase):
    def setUp(self):
        from apps.core.models import Restructuring
        self.entity = _make_entity(category='6')
        self.mfmp = _make_mfmp(self.entity, horizon=3)
        _seed_incomes(self.mfmp)
        _seed_expenses(self.mfmp)
        _seed_debts(self.mfmp)

        import datetime
        self.restructuring = Restructuring.objects.create(
            entity=self.entity,
            name='Reestructuracion test',
            reference_date=datetime.date(2026, 1, 1),
        )
        from apps.planta.models import PayrollPlan
        self.plan_empty = PayrollPlan.objects.create(
            entity=self.entity,
            restructuring=self.restructuring,
            name='Plan vacio',
            kind='PROPOSED',
            reference_date=datetime.date(2026, 1, 1),
        )
        self.plan_with_positions = PayrollPlan.objects.create(
            entity=self.entity,
            restructuring=self.restructuring,
            name='Plan con cargos',
            kind='PROPOSED',
            reference_date=datetime.date(2026, 1, 1),
        )
        from apps.planta.models import PayrollPosition
        PayrollPosition.objects.create(
            plan=self.plan_with_positions,
            denomination='Profesional Universitario',
            hierarchy_level='PROFESIONAL',
            quantity=5,
            monthly_salary=Decimal('3800000'),
        )

    def test_simulate_plan_empty_equals_baseline(self):
        """Plan vacío no cambia indicadores respecto al baseline."""
        result = simulate_plan_impact(self.mfmp, self.plan_empty)
        self.assertEqual(result['annual_cost'], 0.0)
        # Con annual_cost = 0, simulated debe ser igual a baseline
        for yr_str in result['baseline']['law_617']:
            base = result['baseline']['law_617'][yr_str]
            sim = result['simulated']['law_617'][yr_str]
            self.assertAlmostEqual(base['ratio'], sim['ratio'], places=6)

    def test_simulate_plan_with_positions_increases_funcionamiento(self):
        """Plan con posiciones aumenta el costo y puede generar broken_years."""
        result = simulate_plan_impact(self.mfmp, self.plan_with_positions)
        self.assertGreater(result['annual_cost'], 0)
        # El funcionamiento simulado debe ser mayor que el baseline
        yr_str = str(self.mfmp.base_year)
        base_func = result['baseline']['law_617'][yr_str]['funcionamiento']
        sim_func = result['simulated']['law_617'][yr_str]['funcionamiento']
        self.assertGreater(sim_func, base_func)

    def test_simulate_returns_required_keys(self):
        result = simulate_plan_impact(self.mfmp, self.plan_empty)
        for key in ['plan_id', 'annual_cost', 'baseline', 'simulated',
                    'broken_years_617', 'broken_years_358']:
            self.assertIn(key, result)


# ---------------------------------------------------------------------------
# Test importación FUT Excel
# ---------------------------------------------------------------------------

class ImportFutExcelTest(TestCase):
    def setUp(self):
        self.entity = _make_entity()
        self.mfmp = _make_mfmp(self.entity, base_year=2026, horizon=2)

    def _make_workbook(self):
        import openpyxl
        wb = openpyxl.Workbook()
        # Hoja INGRESOS
        ws_inc = wb.active
        ws_inc.title = 'INGRESOS'
        ws_inc.append(['concepto', 'año', 'valor'])
        ws_inc.append(['Ingresos tributarios', 2026, 5000000000])
        ws_inc.append(['transferencias sgp', 2026, 12000000000])
        ws_inc.append(['Ingresos tributarios', 2027, 5200000000])
        ws_inc.append(['transferencias sgp', 2027, 12360000000])

        # Hoja GASTOS
        ws_exp = wb.create_sheet('GASTOS')
        ws_exp.append(['concepto', 'año', 'valor'])
        ws_exp.append(['gastos de personal', 2026, 8000000000])
        ws_exp.append(['gastos generales', 2026, 3500000000])
        ws_exp.append(['gastos de personal', 2027, 8320000000])

        # Hoja DEUDA
        ws_debt = wb.create_sheet('DEUDA')
        ws_debt.append(['año', 'saldo', 'servicio', 'nuevos'])
        ws_debt.append([2026, 4000000000, 500000000, 0])
        ws_debt.append([2027, 3500000000, 500000000, 0])

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def test_import_fut_excel(self):
        """Importar FUT crea los registros correctamente (upserts)."""
        buf = self._make_workbook()
        result = import_fut_excel(buf, self.mfmp)
        self.assertEqual(result['incomes_upserted'], 4)
        self.assertEqual(result['expenses_upserted'], 3)
        self.assertEqual(result['debts_upserted'], 2)
        self.assertIsInstance(result['warnings'], list)

    def test_import_fut_idempotent(self):
        """Segunda importación actualiza sin duplicar."""
        buf1 = self._make_workbook()
        buf2 = self._make_workbook()
        r1 = import_fut_excel(buf1, self.mfmp)
        r2 = import_fut_excel(buf2, self.mfmp)
        # Mismos upserts en ambas ejecuciones (update_or_create)
        self.assertEqual(r1['incomes_upserted'], r2['incomes_upserted'])
        # No se duplican registros en DB
        count = MFMPIncomeProjection.objects.filter(mfmp=self.mfmp).count()
        self.assertLessEqual(count, 4)  # 4 combinaciones únicas (concepto × año × mfmp)

    def test_import_fut_missing_sheet_warns(self):
        """Archivo sin hoja INGRESOS devuelve warning."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'OTROS'
        ws.append(['algo'])
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        result = import_fut_excel(buf, self.mfmp)
        self.assertTrue(any('INGRESOS' in w for w in result['warnings']))


# ---------------------------------------------------------------------------
# Test generación de ficha DOCX
# ---------------------------------------------------------------------------

class FiscalImpactSheetTest(TestCase):
    def setUp(self):
        from apps.core.models import Restructuring
        self.entity = _make_entity(category='6')
        self.mfmp = _make_mfmp(self.entity, horizon=3)
        _seed_incomes(self.mfmp)
        _seed_expenses(self.mfmp)
        _seed_debts(self.mfmp)

        import datetime
        self.restructuring = Restructuring.objects.create(
            entity=self.entity,
            name='Reestructuracion test docx',
            reference_date=datetime.date(2026, 1, 1),
        )
        from apps.planta.models import PayrollPlan
        self.plan = PayrollPlan.objects.create(
            entity=self.entity,
            restructuring=self.restructuring,
            name='Plan vacio test',
            kind='PROPOSED',
            reference_date=datetime.date(2026, 1, 1),
        )

    def test_fiscal_impact_sheet_generates_docx(self):
        """La ficha DOCX se genera y tiene tamaño mínimo razonable."""
        content = generate_fiscal_impact_sheet(self.mfmp, self.plan)
        self.assertIsInstance(content, bytes)
        self.assertGreaterEqual(len(content), 15000, (
            f'DOCX demasiado pequeño: {len(content)} bytes. '
            'Puede que las tablas no se estén generando.'
        ))


# ---------------------------------------------------------------------------
# Tests de API
# ---------------------------------------------------------------------------

class MFMPAPITest(TestCase):
    def setUp(self):
        self.entity = _make_entity(name='Entidad API Test', category='6')
        self.superuser = User.objects.create_superuser('admin_mfmp', 'a@b.com', 'pass')
        self.client = APIClient()
        self.client.force_authenticate(user=self.superuser)

    def test_api_mfmp_crud_as_superuser(self):
        """POST, GET list, GET detalle, GET ley-617 funcionan correctamente."""
        # Crear MFMP
        resp = self.client.post(
            '/api/mfmp/',
            {'entity': self.entity.pk, 'name': 'API MFMP', 'base_year': 2026, 'horizon_years': 3},
            HTTP_X_ENTITY_ID=str(self.entity.pk),
        )
        self.assertEqual(resp.status_code, 201, resp.data)
        mfmp_id = resp.data['id']

        # Listar
        resp = self.client.get('/api/mfmp/', HTTP_X_ENTITY_ID=str(self.entity.pk))
        self.assertEqual(resp.status_code, 200)
        self.assertGreaterEqual(len(resp.data), 1)

        # Detalle
        resp = self.client.get(f'/api/mfmp/{mfmp_id}/', HTTP_X_ENTITY_ID=str(self.entity.pk))
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.data['name'], 'API MFMP')

        # Ley 617 (sin datos = ratios 0)
        resp = self.client.get(f'/api/mfmp/{mfmp_id}/ley-617/', HTTP_X_ENTITY_ID=str(self.entity.pk))
        self.assertEqual(resp.status_code, 200)
        # Debe tener los 3 años del horizonte
        self.assertIn('2026', resp.data)

    def test_api_matriz_action(self):
        """Acción /matriz/ devuelve la estructura correcta."""
        mfmp = _make_mfmp(self.entity, horizon=2)
        resp = self.client.get(
            f'/api/mfmp/{mfmp.pk}/matriz/',
            HTTP_X_ENTITY_ID=str(self.entity.pk),
        )
        self.assertEqual(resp.status_code, 200)
        self.assertIn('years', resp.data)
        self.assertIn('totals', resp.data)


# ---------------------------------------------------------------------------
# Tests de constraints
# ---------------------------------------------------------------------------

class MFMPConstraintTest(TestCase):
    def setUp(self):
        self.entity = _make_entity()
        self.mfmp = _make_mfmp(self.entity, horizon=2)

    def test_scenario_unique_baseline(self):
        """Solo puede haber un escenario baseline por MFMP."""
        from django.db import IntegrityError
        MFMPScenario.objects.create(
            mfmp=self.mfmp, name='Baseline', is_baseline=True
        )
        with self.assertRaises(IntegrityError):
            MFMPScenario.objects.create(
                mfmp=self.mfmp, name='Otro baseline', is_baseline=True
            )

    def test_unique_together_income(self):
        """No puede haber dos ingresos del mismo (mfmp, year, concept)."""
        from django.db import IntegrityError
        MFMPIncomeProjection.objects.create(
            mfmp=self.mfmp, year=2026,
            concept=IncomeConcept.TRIBUTARIOS,
            amount=Decimal('1000000'),
        )
        with self.assertRaises(IntegrityError):
            MFMPIncomeProjection.objects.create(
                mfmp=self.mfmp, year=2026,
                concept=IncomeConcept.TRIBUTARIOS,
                amount=Decimal('2000000'),
            )


# ---------------------------------------------------------------------------
# Tests de aislamiento multi-tenant
# ---------------------------------------------------------------------------

class MFMPEntityIsolationTest(TestCase):
    def tearDown(self):
        # Limpiar el cache de permisos para no interferir con otros tests
        from django.core.cache import cache
        cache.clear()

    def setUp(self):
        from django.contrib.auth.models import Group
        from apps.core.models import GroupModelPermission

        self.entity_a = _make_entity(name='Entidad Iso A')
        self.entity_b = _make_entity(name='Entidad Iso B')

        # Crear un grupo con permiso de lectura en mfmp.MFMP
        self.group = Group.objects.create(name='mfmp_readers')
        GroupModelPermission.objects.create(
            group=self.group,
            app_label='mfmp',
            model='mfmp',
            can_create=False, can_read=True, can_update=False, can_delete=False,
        )

        # user_a con acceso solo a entity_a y en el grupo con permisos
        self.user_a = User.objects.create_user('user_iso_a', password='pass_a')
        self.user_a.groups.add(self.group)
        UserEntityAccess.objects.create(user=self.user_a, entity=self.entity_a, is_default=True)

        self.mfmp_b = _make_mfmp(self.entity_b, horizon=2)

        self.client = APIClient()
        self.client.force_authenticate(user=self.user_a)

    def test_entity_isolation(self):
        """Usuario con acceso a A recibe 403 entity_not_authorized al usar entity_b."""
        resp = self.client.get(
            '/api/mfmp/',
            HTTP_X_ENTITY_ID=str(self.entity_b.pk),
        )
        # 403 entity_not_authorized
        self.assertEqual(resp.status_code, 403)
        self.assertIn('entity_not_authorized', str(resp.data))

    def test_entity_a_only_sees_own_mfmps(self):
        """Superuser con entity_a ve solo los MFMP de esa entidad (queryset isolation)."""
        # Usar superuser para pasar MatrixPermission sin restricciones de grupo
        superuser = User.objects.create_superuser('iso_super', 'iso@b.com', 'pass')
        client = APIClient()
        client.force_authenticate(user=superuser)

        mfmp_a = _make_mfmp(self.entity_a, horizon=2)
        resp = client.get(
            '/api/mfmp/',
            HTTP_X_ENTITY_ID=str(self.entity_a.pk),
        )
        self.assertEqual(resp.status_code, 200)
        # La respuesta puede estar paginada
        results = resp.data.get('results', resp.data) if isinstance(resp.data, dict) else resp.data
        ids = [item['id'] for item in results]
        self.assertIn(mfmp_a.pk, ids)
        self.assertNotIn(self.mfmp_b.pk, ids)
