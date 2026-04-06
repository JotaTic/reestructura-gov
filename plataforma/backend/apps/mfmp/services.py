"""
Servicios de negocio para el módulo MFMP (M17) — Ley 819/2003.

Cálculos de proyección fiscal, indicadores Ley 617/2000, Ley 358/1997
y simulación de impacto de planta de personal.
"""
from __future__ import annotations

from decimal import Decimal
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from apps.mfmp.models import MFMP
    from apps.planta.models import PayrollPlan


# ---------------------------------------------------------------------------
# Límites Ley 617/2000 por categoría municipal
# ---------------------------------------------------------------------------

_LAW_617_LIMITS: dict[str, Decimal] = {
    'ESPECIAL': Decimal('0.50'),
    '1': Decimal('0.65'),
    '2': Decimal('0.70'),
    '3': Decimal('0.70'),
    '4': Decimal('0.80'),
    '5': Decimal('0.80'),
    '6': Decimal('0.80'),
    'NA': Decimal('0.80'),
}

# Órdenes no municipales (simplificado: 50%)
_NON_MUNICIPAL_LIMIT = Decimal('0.50')


def _get_617_limit(entity) -> Decimal:
    order = getattr(entity, 'order', 'MUNICIPAL')
    if order in ('NACIONAL', 'DEPARTAMENTAL', 'DISTRITAL'):
        return _NON_MUNICIPAL_LIMIT
    cat = getattr(entity, 'municipality_category', 'NA') or 'NA'
    return _LAW_617_LIMITS.get(cat, Decimal('0.80'))


# ---------------------------------------------------------------------------
# Matriz de proyección
# ---------------------------------------------------------------------------

def get_projection_matrix(mfmp: 'MFMP') -> dict:
    """
    Devuelve la matriz de proyección del MFMP con todas las cifras organizadas.

    Retorna:
    {
        'years': [...],
        'income_by_concept': {concept: {year: amount}},
        'expense_by_concept': {concept: {year: amount}},
        'debt': {year: {outstanding_debt, debt_service, new_disbursements}},
        'totals': {
            'income': {year: total},
            'expense': {year: total},
        }
    }
    """
    years = list(range(mfmp.base_year, mfmp.base_year + mfmp.horizon_years))

    income_by_concept: dict[str, dict[int, Decimal]] = {}
    for inc in mfmp.incomes.all():
        income_by_concept.setdefault(inc.concept, {})[inc.year] = inc.amount

    expense_by_concept: dict[str, dict[int, Decimal]] = {}
    for exp in mfmp.expenses.all():
        expense_by_concept.setdefault(exp.concept, {})[exp.year] = exp.amount

    debt: dict[int, dict] = {}
    for d in mfmp.debts.all():
        debt[d.year] = {
            'outstanding_debt': d.outstanding_debt,
            'debt_service': d.debt_service,
            'new_disbursements': d.new_disbursements,
        }

    income_totals: dict[int, Decimal] = {}
    for year in years:
        total = sum(
            concept_data.get(year, Decimal('0'))
            for concept_data in income_by_concept.values()
        )
        income_totals[year] = total

    expense_totals: dict[int, Decimal] = {}
    for year in years:
        total = sum(
            concept_data.get(year, Decimal('0'))
            for concept_data in expense_by_concept.values()
        )
        expense_totals[year] = total

    return {
        'years': years,
        'income_by_concept': {
            k: {str(yr): float(v) for yr, v in vd.items()}
            for k, vd in income_by_concept.items()
        },
        'expense_by_concept': {
            k: {str(yr): float(v) for yr, v in vd.items()}
            for k, vd in expense_by_concept.items()
        },
        'debt': {
            str(yr): {
                'outstanding_debt': float(d['outstanding_debt']),
                'debt_service': float(d['debt_service']),
                'new_disbursements': float(d['new_disbursements']),
            }
            for yr, d in debt.items()
        },
        'totals': {
            'income': {str(yr): float(v) for yr, v in income_totals.items()},
            'expense': {str(yr): float(v) for yr, v in expense_totals.items()},
        },
    }


# ---------------------------------------------------------------------------
# Ley 617/2000 — Límites de funcionamiento sobre ICLD
# ---------------------------------------------------------------------------

def _compute_617(mfmp: 'MFMP', personal_deltas: dict[int, Decimal] | None = None) -> dict:
    """
    Calcula los indicadores de Ley 617 por año, con soporte para deltas.

    personal_deltas: {year: delta_amount} para simular incrementos en
    FUNCIONAMIENTO_PERSONAL. No modifica la base de datos.
    """
    years = list(range(mfmp.base_year, mfmp.base_year + mfmp.horizon_years))
    limit = _get_617_limit(mfmp.entity)
    result = {}

    # Agrupar ingresos y gastos por año
    income_by_year: dict[int, dict[str, Decimal]] = {yr: {} for yr in years}
    for inc in mfmp.incomes.all():
        if inc.year in income_by_year:
            income_by_year[inc.year][inc.concept] = inc.amount

    expense_by_year: dict[int, dict[str, Decimal]] = {yr: {} for yr in years}
    for exp in mfmp.expenses.all():
        if exp.year in expense_by_year:
            expense_by_year[exp.year][exp.concept] = exp.amount

    for yr in years:
        inc = income_by_year[yr]
        exp = expense_by_year[yr]

        # ICLD = Tributarios + No tributarios (ingresos corrientes de libre destinación)
        icld = (
            inc.get('TRIBUTARIOS', Decimal('0'))
            + inc.get('NO_TRIBUTARIOS', Decimal('0'))
        )

        # Gastos de funcionamiento
        personal = exp.get('FUNCIONAMIENTO_PERSONAL', Decimal('0'))
        # Aplicar delta si existe
        if personal_deltas and yr in personal_deltas:
            personal += personal_deltas[yr]

        funcionamiento = (
            personal
            + exp.get('FUNCIONAMIENTO_GENERALES', Decimal('0'))
            + exp.get('FUNCIONAMIENTO_TRANSFERENCIAS', Decimal('0'))
        )

        if icld > 0:
            ratio = (funcionamiento / icld).quantize(Decimal('0.0001'))
        else:
            ratio = Decimal('0')

        result[yr] = {
            'icld': float(icld),
            'funcionamiento': float(funcionamiento),
            'ratio': float(ratio),
            'limit': float(limit),
            'compliant': ratio <= limit,
        }

    return result


def calculate_law_617_by_year(mfmp: 'MFMP') -> dict:
    """
    Ley 617/2000: gastos funcionamiento / ICLD <= límite por categoría.

    ICLD = TRIBUTARIOS + NO_TRIBUTARIOS.
    Gastos funcionamiento = FUNCIONAMIENTO_PERSONAL + FUNCIONAMIENTO_GENERALES +
                            FUNCIONAMIENTO_TRANSFERENCIAS.
    Límite por categoría municipal.

    Devuelve {year: {icld, funcionamiento, ratio, limit, compliant}}.
    """
    return _compute_617(mfmp)


# ---------------------------------------------------------------------------
# Ley 358/1997 — Sostenibilidad e indicador de solvencia
# ---------------------------------------------------------------------------

def _compute_358(mfmp: 'MFMP', personal_deltas: dict[int, Decimal] | None = None) -> dict:
    """
    Calcula indicadores Ley 358 por año, con soporte para deltas.
    """
    years = list(range(mfmp.base_year, mfmp.base_year + mfmp.horizon_years))
    result = {}

    income_by_year: dict[int, dict[str, Decimal]] = {yr: {} for yr in years}
    for inc in mfmp.incomes.all():
        if inc.year in income_by_year:
            income_by_year[inc.year][inc.concept] = inc.amount

    expense_by_year: dict[int, dict[str, Decimal]] = {yr: {} for yr in years}
    for exp in mfmp.expenses.all():
        if exp.year in expense_by_year:
            expense_by_year[exp.year][exp.concept] = exp.amount

    debt_by_year: dict[int, dict] = {}
    for d in mfmp.debts.all():
        if d.year in set(years):
            debt_by_year[d.year] = {
                'outstanding_debt': d.outstanding_debt,
                'debt_service': d.debt_service,
            }

    for yr in years:
        inc = income_by_year[yr]
        exp = expense_by_year[yr]
        debt = debt_by_year.get(yr, {})

        # Ingresos corrientes = TRIBUTARIOS + NO_TRIBUTARIOS + TRANSFERENCIAS_SGP +
        #                       TRANSFERENCIAS_OTRAS + REGALIAS
        ingresos_corrientes = (
            inc.get('TRIBUTARIOS', Decimal('0'))
            + inc.get('NO_TRIBUTARIOS', Decimal('0'))
            + inc.get('TRANSFERENCIAS_SGP', Decimal('0'))
            + inc.get('TRANSFERENCIAS_OTRAS', Decimal('0'))
            + inc.get('REGALIAS', Decimal('0'))
        )

        # Gastos de funcionamiento
        personal = exp.get('FUNCIONAMIENTO_PERSONAL', Decimal('0'))
        if personal_deltas and yr in personal_deltas:
            personal += personal_deltas[yr]

        gastos_func = (
            personal
            + exp.get('FUNCIONAMIENTO_GENERALES', Decimal('0'))
            + exp.get('FUNCIONAMIENTO_TRANSFERENCIAS', Decimal('0'))
        )

        # Ahorro operacional = ingresos corrientes - gastos funcionamiento
        ahorro_operacional = ingresos_corrientes - gastos_func

        debt_service = Decimal(str(debt.get('debt_service', 0)))
        outstanding_debt = Decimal(str(debt.get('outstanding_debt', 0)))

        # Solvencia = servicio_deuda / ahorro_operacional
        if ahorro_operacional > 0:
            solvency_ratio = (debt_service / ahorro_operacional).quantize(Decimal('0.0001'))
        else:
            solvency_ratio = Decimal('0')

        # Sostenibilidad = saldo_deuda / ingresos_corrientes
        if ingresos_corrientes > 0:
            sustainability_ratio = (outstanding_debt / ingresos_corrientes).quantize(Decimal('0.0001'))
        else:
            sustainability_ratio = Decimal('0')

        # Semáforo solvencia
        if solvency_ratio < Decimal('0.40'):
            status = 'VERDE'
        elif solvency_ratio < Decimal('0.60'):
            status = 'AMARILLO'
        else:
            status = 'ROJO'

        result[yr] = {
            'solvency_ratio': float(solvency_ratio),
            'sustainability_ratio': float(sustainability_ratio),
            'status': status,
            'ahorro_operacional': float(ahorro_operacional),
            'ingresos_corrientes': float(ingresos_corrientes),
            'debt_service': float(debt_service),
            'outstanding_debt': float(outstanding_debt),
        }

    return result


def calculate_law_358_by_year(mfmp: 'MFMP') -> dict:
    """
    Ley 358/1997: sostenibilidad e indicador de solvencia.

    Solvencia = debt_service / ahorro_operacional.
    Ahorro = ingresos_corrientes - gastos_funcionamiento.
    Sostenibilidad = outstanding_debt / ingresos_corrientes.
    Verde < 40% solvencia, Amarillo 40-60%, Rojo >= 60%.
    Sostenibilidad < 80% es OK.

    Devuelve {year: {solvency_ratio, sustainability_ratio, status, ...}}.
    """
    return _compute_358(mfmp)


# ---------------------------------------------------------------------------
# Simulación de impacto de planta
# ---------------------------------------------------------------------------

def simulate_plan_impact(mfmp: 'MFMP', payroll_plan: 'PayrollPlan') -> dict:
    """
    Simula el impacto fiscal de una planta de personal sobre el MFMP.

    Toma el costo anual efectivo del plan (calculate_payroll_total) e inyecta
    ese valor como DELTA en FUNCIONAMIENTO_PERSONAL para cada año del horizonte.
    Recalcula Ley 617 y Ley 358 con los deltas aplicados.
    NO modifica la DB.

    Retorna:
    {
        'plan_id': ...,
        'annual_cost': ...,
        'baseline': {'law_617': {...}, 'law_358': {...}},
        'simulated': {'law_617': {...}, 'law_358': {...}},
        'broken_years_617': [year, ...],
        'broken_years_358': [year, ...],
    }
    """
    from apps.nomina.services import calculate_payroll_total

    totals = calculate_payroll_total(payroll_plan)
    annual_cost = Decimal(str(totals.get('annual_effective', 0)))

    # Baseline sin deltas
    baseline_617 = _compute_617(mfmp)
    baseline_358 = _compute_358(mfmp)

    # Deltas: igual para cada año del horizonte (costo anual del plan)
    years = list(range(mfmp.base_year, mfmp.base_year + mfmp.horizon_years))
    personal_deltas = {yr: annual_cost for yr in years}

    # Simulado con deltas
    simulated_617 = _compute_617(mfmp, personal_deltas=personal_deltas)
    simulated_358 = _compute_358(mfmp, personal_deltas=personal_deltas)

    # Años que rompen la Ley 617 en simulación (y no en baseline)
    broken_years_617 = [
        yr for yr in years
        if not simulated_617[yr]['compliant']
    ]

    # Años en ROJO en simulación
    broken_years_358 = [
        yr for yr in years
        if simulated_358[yr]['status'] == 'ROJO'
    ]

    return {
        'plan_id': payroll_plan.pk,
        'annual_cost': float(annual_cost),
        'baseline': {
            'law_617': {str(k): v for k, v in baseline_617.items()},
            'law_358': {str(k): v for k, v in baseline_358.items()},
        },
        'simulated': {
            'law_617': {str(k): v for k, v in simulated_617.items()},
            'law_358': {str(k): v for k, v in simulated_358.items()},
        },
        'broken_years_617': broken_years_617,
        'broken_years_358': broken_years_358,
    }


# ---------------------------------------------------------------------------
# Importación de FUT (Formato Único Territorial) desde Excel
# ---------------------------------------------------------------------------

# Mapeo de palabras clave a conceptos de ingreso
_INCOME_KEYWORD_MAP: list[tuple[str, str]] = [
    ('tribut', 'TRIBUTARIOS'),
    ('no tribut', 'NO_TRIBUTARIOS'),
    ('no_tribut', 'NO_TRIBUTARIOS'),
    ('sgp', 'TRANSFERENCIAS_SGP'),
    ('propósito general', 'TRANSFERENCIAS_SGP'),
    ('proposito general', 'TRANSFERENCIAS_SGP'),
    ('transfer', 'TRANSFERENCIAS_OTRAS'),
    ('regal', 'REGALIAS'),
    ('cofinanci', 'COFINANCIACION'),
    ('credit', 'CREDITO'),
    ('balance', 'RECURSOS_BALANCE'),
]

# Mapeo de palabras clave a conceptos de gasto
_EXPENSE_KEYWORD_MAP: list[tuple[str, str]] = [
    ('personal', 'FUNCIONAMIENTO_PERSONAL'),
    ('generales', 'FUNCIONAMIENTO_GENERALES'),
    ('general', 'FUNCIONAMIENTO_GENERALES'),
    ('transferenci', 'FUNCIONAMIENTO_TRANSFERENCIAS'),
    ('servicio deuda', 'SERVICIO_DEUDA'),
    ('servicio_deuda', 'SERVICIO_DEUDA'),
    ('deuda', 'SERVICIO_DEUDA'),
    ('inversion', 'INVERSION'),
    ('inversión', 'INVERSION'),
]


def _map_concept(raw: str, keyword_map: list[tuple[str, str]], default='OTROS') -> str:
    """Normaliza un concepto libre a uno de los enums usando palabras clave."""
    raw_lower = str(raw).lower().strip()
    for keyword, concept in keyword_map:
        if keyword in raw_lower:
            return concept
    return default


def import_fut_excel(file, mfmp: 'MFMP') -> dict:
    """
    Parser tolerante del FUT (Formato Único Territorial).

    Acepta un Excel con hojas 'INGRESOS' y 'GASTOS' con columnas
    'concepto', 'año', 'valor'.
    Hoja opcional 'DEUDA' con columnas 'año', 'saldo', 'servicio', 'nuevos'.

    Idempotente: update_or_create por (mfmp, year, concept).

    Retorna {'incomes_upserted', 'expenses_upserted', 'debts_upserted', 'warnings'}.
    """
    import openpyxl
    from apps.mfmp.models import MFMPIncomeProjection, MFMPExpenseProjection, MFMPDebtProjection

    try:
        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    except Exception as e:
        return {'incomes_upserted': 0, 'expenses_upserted': 0, 'debts_upserted': 0,
                'warnings': [f'Error al leer el archivo: {e}']}

    warnings: list[str] = []
    incomes_upserted = 0
    expenses_upserted = 0
    debts_upserted = 0

    def _get_sheet(name):
        # Búsqueda insensible a mayúsculas
        for sname in wb.sheetnames:
            if sname.upper() == name.upper():
                return wb[sname]
        return None

    def _col_index(headers_row, *names) -> int | None:
        """Busca el índice de columna (0-based) por varios nombres posibles."""
        cols = [str(c.value or '').lower().strip() for c in headers_row]
        for name in names:
            name_lower = name.lower()
            if name_lower in cols:
                return cols.index(name_lower)
        return None

    # ---- INGRESOS ----
    ws = _get_sheet('INGRESOS')
    if ws is None:
        warnings.append("Hoja 'INGRESOS' no encontrada.")
    else:
        rows = list(ws.iter_rows())
        if len(rows) < 2:
            warnings.append("Hoja 'INGRESOS' sin datos.")
        else:
            headers = rows[0]
            ci_concept = _col_index(headers, 'concepto', 'concept')
            ci_year = _col_index(headers, 'año', 'year', 'anio')
            ci_valor = _col_index(headers, 'valor', 'value', 'monto', 'amount')
            if ci_concept is None or ci_year is None or ci_valor is None:
                warnings.append("Hoja 'INGRESOS': columnas requeridas no encontradas (concepto, año, valor).")
            else:
                for row in rows[1:]:
                    try:
                        raw_concept = row[ci_concept].value
                        raw_year = row[ci_year].value
                        raw_valor = row[ci_valor].value
                        if raw_concept is None and raw_year is None:
                            continue
                        concept = _map_concept(raw_concept or '', _INCOME_KEYWORD_MAP)
                        year = int(raw_year)
                        amount = Decimal(str(raw_valor or 0))
                        MFMPIncomeProjection.objects.update_or_create(
                            mfmp=mfmp, year=year, concept=concept,
                            defaults={'amount': amount},
                        )
                        incomes_upserted += 1
                    except Exception as e:
                        warnings.append(f'INGRESOS fila ignorada: {e}')

    # ---- GASTOS ----
    ws = _get_sheet('GASTOS')
    if ws is None:
        warnings.append("Hoja 'GASTOS' no encontrada.")
    else:
        rows = list(ws.iter_rows())
        if len(rows) < 2:
            warnings.append("Hoja 'GASTOS' sin datos.")
        else:
            headers = rows[0]
            ci_concept = _col_index(headers, 'concepto', 'concept')
            ci_year = _col_index(headers, 'año', 'year', 'anio')
            ci_valor = _col_index(headers, 'valor', 'value', 'monto', 'amount')
            if ci_concept is None or ci_year is None or ci_valor is None:
                warnings.append("Hoja 'GASTOS': columnas requeridas no encontradas (concepto, año, valor).")
            else:
                for row in rows[1:]:
                    try:
                        raw_concept = row[ci_concept].value
                        raw_year = row[ci_year].value
                        raw_valor = row[ci_valor].value
                        if raw_concept is None and raw_year is None:
                            continue
                        concept = _map_concept(raw_concept or '', _EXPENSE_KEYWORD_MAP)
                        year = int(raw_year)
                        amount = Decimal(str(raw_valor or 0))
                        MFMPExpenseProjection.objects.update_or_create(
                            mfmp=mfmp, year=year, concept=concept,
                            defaults={'amount': amount},
                        )
                        expenses_upserted += 1
                    except Exception as e:
                        warnings.append(f'GASTOS fila ignorada: {e}')

    # ---- DEUDA (opcional) ----
    ws = _get_sheet('DEUDA')
    if ws is not None:
        rows = list(ws.iter_rows())
        if len(rows) >= 2:
            headers = rows[0]
            ci_year = _col_index(headers, 'año', 'year', 'anio')
            ci_saldo = _col_index(headers, 'saldo', 'outstanding', 'outstanding_debt')
            ci_servicio = _col_index(headers, 'servicio', 'service', 'debt_service')
            ci_nuevos = _col_index(headers, 'nuevos', 'new', 'new_disbursements', 'desembolsos')
            if ci_year is not None:
                for row in rows[1:]:
                    try:
                        raw_year = row[ci_year].value
                        if raw_year is None:
                            continue
                        year = int(raw_year)
                        saldo = Decimal(str(row[ci_saldo].value or 0)) if ci_saldo is not None else Decimal('0')
                        servicio = Decimal(str(row[ci_servicio].value or 0)) if ci_servicio is not None else Decimal('0')
                        nuevos = Decimal(str(row[ci_nuevos].value or 0)) if ci_nuevos is not None else Decimal('0')
                        from apps.mfmp.models import MFMPDebtProjection
                        MFMPDebtProjection.objects.update_or_create(
                            mfmp=mfmp, year=year,
                            defaults={
                                'outstanding_debt': saldo,
                                'debt_service': servicio,
                                'new_disbursements': nuevos,
                            },
                        )
                        debts_upserted += 1
                    except Exception as e:
                        warnings.append(f'DEUDA fila ignorada: {e}')

    return {
        'incomes_upserted': incomes_upserted,
        'expenses_upserted': expenses_upserted,
        'debts_upserted': debts_upserted,
        'warnings': warnings,
    }


# ---------------------------------------------------------------------------
# Ficha de impacto fiscal (Ley 819 art. 7)
# ---------------------------------------------------------------------------

def generate_fiscal_impact_sheet(mfmp: 'MFMP', plan: 'PayrollPlan') -> bytes:
    """
    Ficha oficial estilo Ley 819 art. 7 como DOCX.
    Reutiliza apps.common.exports.build_docx.

    Retorna bytes del docx para que la vista construya el HttpResponse.
    """
    from apps.common.exports import build_docx

    sim = simulate_plan_impact(mfmp, plan)
    baseline_617 = sim['baseline']['law_617']
    simulated_617 = sim['simulated']['law_617']
    baseline_358 = sim['baseline']['law_358']
    simulated_358 = sim['simulated']['law_358']
    years = sorted(int(y) for y in baseline_617.keys())

    title = f'Ficha de Impacto Fiscal — Ley 819/2003'
    meta = [
        ('Entidad', mfmp.entity.name),
        ('Plan de personal', str(plan)),
        ('MFMP año base', str(mfmp.base_year)),
        ('Horizonte', f'{mfmp.horizon_years} años'),
        ('Aprobado por', mfmp.approved_by or '—'),
        ('Costo anual estimado del plan', f'${sim["annual_cost"]:,.2f}'),
    ]

    sections = []

    # Identificación
    sections.append({
        'heading': 'Identificación',
        'headers': ['Campo', 'Valor'],
        'rows': [
            ['Entidad', mfmp.entity.name],
            ['NIT', getattr(mfmp.entity, 'nit', '') or '—'],
            ['Orden', mfmp.entity.get_order_display()],
            ['Plan de personal', str(plan)],
            ['MFMP', mfmp.name],
            ['Año base', str(mfmp.base_year)],
            ['Horizonte', f'{mfmp.horizon_years} años'],
            ['Costo anual plan', f'${sim["annual_cost"]:,.2f}'],
        ],
    })

    # Baseline Ley 617
    rows_617_base = []
    for yr in years:
        d = baseline_617.get(str(yr), {})
        estado = 'Cumple' if d.get('compliant', True) else 'No cumple'
        rows_617_base.append([
            yr,
            f'${d.get("icld", 0):,.0f}',
            f'${d.get("funcionamiento", 0):,.0f}',
            f'{d.get("ratio", 0)*100:.1f}%',
            f'{d.get("limit", 0)*100:.0f}%',
            estado,
        ])

    sections.append({
        'heading': 'Baseline Ley 617/2000 por año',
        'description': (
            'ICLD = Tributarios + No tributarios. '
            'Funcionamiento = Personal + Generales + Transferencias. '
            'Límite según categoría de la entidad.'
        ),
        'headers': ['Año', 'ICLD', 'Funcionamiento', 'Ratio', 'Límite', 'Estado'],
        'rows': rows_617_base,
        'notes': 'Fuente: Ley 617/2000 art. 6.',
    })

    # Simulación Ley 617
    rows_617_sim = []
    for yr in years:
        base = baseline_617.get(str(yr), {})
        sim_d = simulated_617.get(str(yr), {})
        delta = sim['annual_cost']
        nuevo_func = sim_d.get('funcionamiento', 0)
        nuevo_ratio = sim_d.get('ratio', 0)
        estado = 'Cumple' if sim_d.get('compliant', True) else 'NO CUMPLE'
        rows_617_sim.append([
            yr,
            f'${delta:,.0f}',
            f'${nuevo_func:,.0f}',
            f'{nuevo_ratio*100:.1f}%',
            estado,
        ])

    sections.append({
        'heading': 'Simulación con planta propuesta — Ley 617/2000',
        'description': (
            f'Delta de personal = ${sim["annual_cost"]:,.2f}/año (costo efectivo del plan). '
            'Se inyecta en FUNCIONAMIENTO_PERSONAL de cada año del horizonte.'
        ),
        'headers': ['Año', 'Delta personal', 'Nuevo Funcionamiento', 'Nuevo Ratio', 'Estado'],
        'rows': rows_617_sim,
    })

    # Años que rompen límites
    broken = sim['broken_years_617']
    broken_358 = sim['broken_years_358']
    broken_rows = []
    if broken:
        for yr in broken:
            broken_rows.append([str(yr), 'Ley 617', 'Límite de funcionamiento superado'])
    if broken_358:
        for yr in broken_358:
            broken_rows.append([str(yr), 'Ley 358', 'Indicador de solvencia en ROJO'])
    if not broken_rows:
        broken_rows = [['—', '—', 'Sin años problemáticos en el horizonte.']]

    sections.append({
        'heading': 'Años que rompen límites',
        'headers': ['Año', 'Ley', 'Descripción'],
        'rows': broken_rows,
    })

    # Indicadores Ley 358
    rows_358 = []
    for yr in years:
        base = baseline_358.get(str(yr), {})
        sim_d = simulated_358.get(str(yr), {})
        rows_358.append([
            yr,
            f'{base.get("solvency_ratio", 0)*100:.1f}%',
            base.get('status', '—'),
            f'{sim_d.get("solvency_ratio", 0)*100:.1f}%',
            sim_d.get('status', '—'),
        ])

    sections.append({
        'heading': 'Indicadores Ley 358/1997 — Baseline vs Simulado',
        'description': (
            'Solvencia = Servicio deuda / Ahorro operacional. '
            'Verde < 40%, Amarillo 40-60%, Rojo >= 60%.'
        ),
        'headers': ['Año', 'Solvencia base', 'Estado base', 'Solvencia sim.', 'Estado sim.'],
        'rows': rows_358,
        'notes': 'Fuente: Ley 358/1997.',
    })

    return build_docx(title, meta, sections)
