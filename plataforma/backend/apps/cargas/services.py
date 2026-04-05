"""
Servicios de dominio para el módulo de cargas de trabajo:
- Consolidación (Formulario 2).
- Cálculo de personal requerido por nivel y por cargo.
- Exportación a Excel compatible con el Anexo 5 de Función Pública.
"""
from collections import defaultdict
from decimal import Decimal
from io import BytesIO
from math import ceil

from .models import WorkloadMatrix, HOURS_PER_MONTH


def consolidate_by_level(matrix: WorkloadMatrix) -> dict:
    """
    Consolida horas-hombre/mes y personal requerido por nivel jerárquico y por dependencia.
    Devuelve el Formulario 2.
    """
    by_dept: dict[int, dict] = {}
    totals_by_level: dict[str, Decimal] = defaultdict(lambda: Decimal('0'))

    for entry in matrix.entries.select_related('department').all():
        dept_id = entry.department_id
        if dept_id not in by_dept:
            by_dept[dept_id] = {
                'department_id': dept_id,
                'department_name': entry.department.name,
                'hours': defaultdict(lambda: Decimal('0')),
                'positions': defaultdict(int),
            }
        by_dept[dept_id]['hours'][entry.hierarchy_level] += entry.hh_month
        totals_by_level[entry.hierarchy_level] += entry.hh_month

    # Convertir horas → personas (÷167, redondeo hacia arriba)
    departments = []
    for dept in by_dept.values():
        total_dept = 0
        positions = {}
        for level, hours in dept['hours'].items():
            required = int(ceil(hours / HOURS_PER_MONTH)) if hours > 0 else 0
            positions[level] = required
            total_dept += required
        departments.append({
            'department_id': dept['department_id'],
            'department_name': dept['department_name'],
            'hours_by_level': {k: float(v) for k, v in dept['hours'].items()},
            'positions_by_level': positions,
            'total_positions': total_dept,
        })

    totals_positions = {
        level: int(ceil(h / HOURS_PER_MONTH)) if h > 0 else 0
        for level, h in totals_by_level.items()
    }
    return {
        'matrix_id': matrix.id,
        'matrix_name': matrix.name,
        'entity': matrix.entity.name,
        'departments': departments,
        'totals': {
            'hours_by_level': {k: float(v) for k, v in totals_by_level.items()},
            'positions_by_level': totals_positions,
            'total_positions': sum(totals_positions.values()),
        },
    }


def consolidate_by_job(matrix: WorkloadMatrix) -> list[dict]:
    """
    Consolida por cargo específico (denominación + código + grado) — útil para el
    comparativo planta actual vs. propuesta y para alimentar el manual de funciones.
    """
    acc: dict[tuple, dict] = {}
    for entry in matrix.entries.select_related('department').all():
        key = (entry.job_denomination, entry.job_code, entry.job_grade, entry.department_id)
        if key not in acc:
            acc[key] = {
                'department_id': entry.department_id,
                'department_name': entry.department.name,
                'job_denomination': entry.job_denomination,
                'job_code': entry.job_code,
                'job_grade': entry.job_grade,
                'hierarchy_level': entry.hierarchy_level,
                'main_purpose': entry.main_purpose,
                'requirements': entry.requirements,
                'total_hours': Decimal('0'),
                'activities': [],
            }
        acc[key]['total_hours'] += entry.hh_month
        acc[key]['activities'].append(entry.activity)
        if entry.requirements and not acc[key]['requirements']:
            acc[key]['requirements'] = entry.requirements

    result = []
    for row in acc.values():
        hours = row['total_hours']
        positions = int(ceil(hours / HOURS_PER_MONTH)) if hours > 0 else 0
        row['total_hours'] = float(hours)
        row['positions_required'] = positions
        result.append(row)
    return result


def build_functions_manual(matrix: WorkloadMatrix) -> dict:
    """
    Manual específico de funciones y competencias laborales (Módulo 12).

    Agrupa por cargo (denominación + código + grado), consolidando propósito,
    funciones (derivadas de actividades medidas) y requisitos.
    """
    from collections import OrderedDict

    grouped: dict[tuple, dict] = OrderedDict()
    for row in consolidate_by_job(matrix):
        key = (row['job_denomination'], row['job_code'], row['job_grade'])
        if key not in grouped:
            grouped[key] = {
                'job_denomination': row['job_denomination'],
                'job_code': row['job_code'],
                'job_grade': row['job_grade'],
                'hierarchy_level': row['hierarchy_level'],
                'main_purpose': row['main_purpose'] or '',
                'requirements': row['requirements'] or '',
                'functions': [],
                'departments': set(),
                'positions_required': 0,
            }
        # Acumular funciones únicas
        for act in row['activities']:
            if act and act not in grouped[key]['functions']:
                grouped[key]['functions'].append(act)
        grouped[key]['departments'].add(row['department_name'])
        grouped[key]['positions_required'] += row['positions_required']

    jobs = []
    for row in grouped.values():
        row['departments'] = sorted(row['departments'])
        jobs.append(row)

    return {
        'matrix_id': matrix.id,
        'matrix_name': matrix.name,
        'entity': matrix.entity.name,
        'nomenclature_decree': matrix.entity.nomenclature_decree,
        'jobs': jobs,
    }


def export_to_xlsx(matrix: WorkloadMatrix) -> bytes:
    """
    Exporta la matriz a un archivo .xlsx con el formato del Anexo 5:
    - Una hoja por dependencia (Formulario 1).
    - Una hoja RESUMEN EMPLEOS ENTIDAD (Formulario 2).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    wb.remove(wb.active)

    headers = [
        '1 Proceso', '2 Actividad', '3 Procedimiento', '4 Nivel jerárquico',
        '5 Requisitos', '6 Denominación', '7 Código', '8 Grado', '9 Propósito principal',
        '10 Frecuencia/mes', '11 Tmin', '12 TU', '13 Tmax', '14 TE',
        '15 Asesor', '16 Profesional', '17 Técnico', '18 Asistencial',
    ]
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='0E7490')

    # Agrupar por dependencia
    depts: dict[int, list] = defaultdict(list)
    for entry in matrix.entries.select_related('department').all():
        depts[entry.department_id].append(entry)

    for dept_id, entries in depts.items():
        dept_name = entries[0].department.name[:28] or f'Dep {dept_id}'
        ws = wb.create_sheet(title=dept_name[:31])
        ws.append([f'ENTIDAD: {matrix.entity.name}'])
        ws.append([f'DEPENDENCIA: {entries[0].department.name}'])
        ws.append([f'FECHA: {matrix.reference_date.isoformat()}'])
        ws.append([])
        ws.append(headers)
        for c in ws[5]:
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal='center', wrap_text=True)

        for e in entries:
            row = [
                e.process, e.activity, e.procedure, e.get_hierarchy_level_display(),
                e.requirements, e.job_denomination, e.job_code, e.job_grade, e.main_purpose,
                float(e.monthly_frequency), float(e.t_min), float(e.t_usual), float(e.t_max),
                float(e.standard_time),
                float(e.hh_month) if e.hierarchy_level == 'ASESOR' else '',
                float(e.hh_month) if e.hierarchy_level == 'PROFESIONAL' else '',
                float(e.hh_month) if e.hierarchy_level == 'TECNICO' else '',
                float(e.hh_month) if e.hierarchy_level == 'ASISTENCIAL' else '',
            ]
            ws.append(row)

        for col_letter, width in zip('ABCDEFGHIJKLMNOPQR', [18, 40, 30, 14, 40, 24, 8, 8, 40, 12, 10, 10, 10, 12, 12, 14, 12, 14]):
            ws.column_dimensions[col_letter].width = width

    # Hoja consolidado (Formulario 2)
    summary = consolidate_by_level(matrix)
    ws = wb.create_sheet(title='RESUMEN EMPLEOS ENTIDAD')
    ws.append([f'ENTIDAD: {summary["entity"]}'])
    ws.append([f'MATRIZ: {summary["matrix_name"]}'])
    ws.append([])
    ws.append(['DEPENDENCIA', 'DIRECTIVO', 'ASESOR', 'PROFESIONAL', 'TÉCNICO', 'ASISTENCIAL', 'TOTAL'])
    for c in ws[4]:
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center')
    for dept in summary['departments']:
        p = dept['positions_by_level']
        ws.append([
            dept['department_name'],
            p.get('DIRECTIVO', 0),
            p.get('ASESOR', 0),
            p.get('PROFESIONAL', 0),
            p.get('TECNICO', 0),
            p.get('ASISTENCIAL', 0),
            dept['total_positions'],
        ])
    totals = summary['totals']['positions_by_level']
    ws.append([
        'TOTAL ENTIDAD',
        totals.get('DIRECTIVO', 0),
        totals.get('ASESOR', 0),
        totals.get('PROFESIONAL', 0),
        totals.get('TECNICO', 0),
        totals.get('ASISTENCIAL', 0),
        summary['totals']['total_positions'],
    ])
    for col_letter, width in zip('ABCDEFG', [45, 12, 12, 14, 12, 14, 12]):
        ws.column_dimensions[col_letter].width = width

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
