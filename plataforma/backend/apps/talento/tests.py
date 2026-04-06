"""
Tests para el módulo de Hojas de vida (M15).

Cubre:
- import_sigep_excel con Workbook en memoria.
- calculate_retirement_eligibility para pre-pensionado.
- GET /api/empleados/ con usuario Consulta (200) y POST (403).
- GET /api/objetivos/definitions/ devuelve 16 kinds.
"""
from datetime import date
from decimal import Decimal
from io import BytesIO

from django.contrib.auth.models import User, Group
from django.test import TestCase
from rest_framework.test import APIClient

from apps.core.models import Entity, Restructuring, UserEntityAccess
from apps.talento.models import Employee
from apps.talento.services import (
    import_sigep_excel,
    calculate_retirement_eligibility,
    calculate_tenure,
)


def _make_entity(name='Test Entity', order='MUNICIPAL'):
    return Entity.objects.create(
        name=name,
        acronym='TE',
        order=order,
        legal_nature='ALCALDIA',
        creation_norm='',
    )


def _make_employee(entity, id_number='99001', sex='M', birth_date=None) -> Employee:
    bd = birth_date or date(1985, 1, 1)
    return Employee.objects.create(
        entity=entity,
        id_type='CC',
        id_number=id_number,
        full_name='Test User',
        first_name='Test',
        last_name='User',
        birth_date=bd,
        sex=sex,
    )


class ImportSigepExcelTest(TestCase):
    """Tests para import_sigep_excel."""

    def setUp(self):
        self.entity = _make_entity()

    def test_imports_two_employees_from_memory_workbook(self):
        """Workbook en memoria con 2 empleados → crea 2 Employee."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Información Básica'
        ws.append(['tipo_doc', 'numero_doc', 'nombres', 'apellidos', 'fecha_nacimiento', 'sexo'])
        ws.append(['CC', '11111111', 'Juan', 'Pérez', '1990-01-15', 'M'])
        ws.append(['CC', '22222222', 'María', 'López', '1985-05-20', 'F'])

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        buf.name = 'test.xlsx'

        result = import_sigep_excel(buf, self.entity)
        self.assertEqual(result['created'], 2)
        self.assertEqual(result['updated'], 0)
        self.assertEqual(Employee.objects.filter(entity=self.entity).count(), 2)

    def test_idempotent_on_second_run(self):
        """Correr dos veces la importación → actualiza en vez de duplicar."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Información Básica'
        ws.append(['tipo_doc', 'numero_doc', 'nombres', 'apellidos', 'fecha_nacimiento', 'sexo'])
        ws.append(['CC', '33333333', 'Pedro', 'García', '1975-03-10', 'M'])

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        buf.name = 'test.xlsx'

        import_sigep_excel(buf, self.entity)
        buf.seek(0)
        result2 = import_sigep_excel(buf, self.entity)
        self.assertEqual(result2['created'], 0)
        self.assertEqual(result2['updated'], 1)
        self.assertEqual(Employee.objects.filter(entity=self.entity, id_number='33333333').count(), 1)

    def test_missing_sheet_reports_warning(self):
        """Si la hoja Información Básica no existe → warning, no error fatal."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'OtraHoja'

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        buf.name = 'test.xlsx'

        result = import_sigep_excel(buf, self.entity)
        self.assertEqual(result['created'], 0)
        self.assertTrue(len(result['warnings']) > 0)


class RetirementEligibilityTest(TestCase):
    """Tests para calculate_retirement_eligibility."""

    def setUp(self):
        self.entity = _make_entity()

    def test_female_56_with_24_years_is_pre_pensioned(self):
        """
        Mujer de 56 años con 24 años de experiencia pública
        → pre-pensionado True (faltan ≤3 años para los 57).
        """
        from apps.talento.models import EmployeeExperience
        emp = _make_employee(
            self.entity,
            id_number='56001',
            sex='F',
            birth_date=date(1969, 1, 1),  # ~56 años en 2026
        )
        EmployeeExperience.objects.create(
            employee=emp,
            employer='Entidad Pública',
            position_name='Profesional',
            sector='PUBLICO',
            start_date=date(2001, 1, 1),
            end_date=date(2025, 1, 1),
            is_public_sector=True,
        )
        result = calculate_retirement_eligibility(emp)
        self.assertTrue(result['is_pre_pensioned'])

    def test_young_male_not_pre_pensioned(self):
        """Hombre joven → no es pre-pensionado."""
        emp = _make_employee(
            self.entity,
            id_number='56002',
            sex='M',
            birth_date=date(1995, 6, 1),  # ~30 años
        )
        result = calculate_retirement_eligibility(emp)
        self.assertFalse(result['is_pre_pensioned'])
        self.assertGreater(result['years_remaining'], 3.0)

    def test_returns_correct_keys(self):
        """Siempre retorna las claves esperadas."""
        emp = _make_employee(self.entity, id_number='56003')
        result = calculate_retirement_eligibility(emp)
        self.assertIn('is_pre_pensioned', result)
        self.assertIn('years_remaining', result)
        self.assertIn('reason', result)


class EmployeeAPITest(TestCase):
    """Tests de API para /api/empleados/."""

    def setUp(self):
        self.entity = _make_entity()

        # Grupo Consulta (read-only)
        consulta_group, _ = Group.objects.get_or_create(name='Consulta')

        self.consulta_user = User.objects.create_user(
            username='testconsulta',
            password='pass1234',
        )
        self.consulta_user.groups.add(consulta_group)
        UserEntityAccess.objects.create(
            user=self.consulta_user,
            entity=self.entity,
            is_default=True,
        )

        # Sembrar un empleado
        _make_employee(self.entity, id_number='55001')

        self.client = APIClient()

    def _login(self, user):
        self.client.force_login(user)

    def test_consulta_can_get_employee_list(self):
        """Usuario del grupo Consulta puede hacer GET /api/empleados/ → 200."""
        from apps.core.models import GroupModelPermission
        # Asegurar permiso de lectura para Consulta en talento.employee
        consulta_group = Group.objects.get(name='Consulta')
        GroupModelPermission.objects.update_or_create(
            group=consulta_group,
            app_label='talento',
            model='employee',
            defaults={'can_read': True, 'can_create': False, 'can_update': False, 'can_delete': False},
        )
        self._login(self.consulta_user)
        resp = self.client.get(
            '/api/empleados/',
            HTTP_X_ENTITY_ID=str(self.entity.id),
        )
        self.assertEqual(resp.status_code, 200)

    def test_consulta_cannot_post_employee(self):
        """Usuario Consulta no puede crear empleado → 403."""
        from apps.core.models import GroupModelPermission
        consulta_group = Group.objects.get(name='Consulta')
        GroupModelPermission.objects.update_or_create(
            group=consulta_group,
            app_label='talento',
            model='employee',
            defaults={'can_read': True, 'can_create': False, 'can_update': False, 'can_delete': False},
        )
        self._login(self.consulta_user)
        resp = self.client.post(
            '/api/empleados/',
            data={
                'entity': self.entity.id,
                'id_type': 'CC',
                'id_number': '99999',
                'full_name': 'Test',
                'first_name': 'Test',
                'last_name': 'User',
                'birth_date': '1990-01-01',
                'sex': 'M',
            },
            format='json',
            HTTP_X_ENTITY_ID=str(self.entity.id),
        )
        self.assertEqual(resp.status_code, 403)


class ObjectiveDefinitionsAPITest(TestCase):
    """Tests para GET /api/objetivos/definitions/."""

    def setUp(self):
        self.entity = _make_entity()
        self.restructuring = Restructuring.objects.create(
            entity=self.entity,
            name='Test Restructuring',
            reference_date=date(2026, 1, 1),
        )
        self.user = User.objects.create_superuser(
            username='supertest',
            password='pass1234',
        )
        self.client = APIClient()
        self.client.force_login(self.user)

    def test_definitions_returns_16_kinds(self):
        """GET /api/objetivos/definitions/ devuelve exactamente 16 kinds."""
        resp = self.client.get(
            '/api/objetivos/definitions/',
            HTTP_X_ENTITY_ID=str(self.entity.id),
            HTTP_X_RESTRUCTURING_ID=str(self.restructuring.id),
        )
        self.assertEqual(resp.status_code, 200)
        data = resp.json()
        self.assertEqual(len(data), 16)

    def test_all_definitions_have_required_keys(self):
        """Cada definición tiene label, required_inputs, active_modules, validators, required_outputs."""
        resp = self.client.get(
            '/api/objetivos/definitions/',
            HTTP_X_ENTITY_ID=str(self.entity.id),
            HTTP_X_RESTRUCTURING_ID=str(self.restructuring.id),
        )
        data = resp.json()
        for kind, defn in data.items():
            with self.subTest(kind=kind):
                self.assertIn('label', defn)
                self.assertIn('required_inputs', defn)
                self.assertIn('active_modules', defn)
                self.assertIn('validators', defn)
                self.assertIn('required_outputs', defn)
