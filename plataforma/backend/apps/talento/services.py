"""
Servicios de negocio para el módulo de Hojas de vida (M15).

- import_sigep_excel: importa empleados desde un archivo Excel de SIGEP.
- calculate_tenure: calcula tiempo de servicio total del empleado.
- calculate_retirement_eligibility: determina si el empleado es pre-pensionado.
"""
from __future__ import annotations

import unicodedata
from datetime import date, timedelta
from decimal import Decimal
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from apps.talento.models import Employee


def _normalize_col(name: str) -> str:
    """Normaliza nombre de columna: minúsculas, sin tildes, sin espacios."""
    if not name:
        return ''
    nfkd = unicodedata.normalize('NFKD', str(name))
    ascii_str = nfkd.encode('ASCII', 'ignore').decode('ASCII')
    return ascii_str.lower().strip().replace(' ', '_').replace('-', '_')


def _get_row_value(row_dict: dict, *possible_keys: str, default=''):
    """Busca el valor de una celda normalizando nombres de columna."""
    for key in possible_keys:
        norm = _normalize_col(key)
        for col_key, val in row_dict.items():
            if _normalize_col(col_key) == norm:
                return val if val is not None else default
    return default


def import_sigep_excel(file, entity) -> dict:
    """
    Importa empleados desde un archivo Excel de SIGEP.

    Hojas esperadas:
    - "Información Básica": tipo_doc, numero_doc, nombres, apellidos,
      fecha_nacimiento, sexo (mínimos). Resto opcionales.
    - "Educación": numero_doc, nivel, institucion, programa, titulo,
      fecha_grado (opcional)
    - "Experiencia": numero_doc, empleador, cargo, sector, fecha_inicio,
      fecha_fin (opcional), es_sector_publico
    - "Capacitación": numero_doc, tema, horas, institucion, fecha

    Retorna: {created: int, updated: int, errors: list[str], warnings: list[str]}
    """
    import openpyxl
    from apps.talento.models import (
        Employee, EmployeeEducation, EmployeeExperience, EmployeeTraining,
    )

    result = {'created': 0, 'updated': 0, 'errors': [], 'warnings': []}

    try:
        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    except Exception as exc:
        result['errors'].append(f'No se pudo abrir el archivo: {exc}')
        return result

    # ---- Hoja: Información Básica ----
    basic_sheet = None
    for sheet_name in wb.sheetnames:
        norm = _normalize_col(sheet_name)
        if 'informacion' in norm and 'basica' in norm or norm in ('basica', 'datos_basicos', 'informacion_basica'):
            basic_sheet = wb[sheet_name]
            break
        if 'basica' in norm or norm == 'empleados':
            basic_sheet = wb[sheet_name]
            break

    if basic_sheet is None:
        result['warnings'].append('Hoja "Información Básica" no encontrada; se omite importación básica.')
    else:
        rows = list(basic_sheet.rows)
        if len(rows) < 2:
            result['warnings'].append('La hoja "Información Básica" no tiene datos.')
        else:
            headers = [str(c.value) if c.value else '' for c in rows[0]]
            for row_idx, row in enumerate(rows[1:], start=2):
                row_dict = {headers[i]: (cell.value if cell.value is not None else '') for i, cell in enumerate(row)}
                # Campos mínimos
                id_type_raw = str(_get_row_value(row_dict, 'tipo_doc', 'tipo de documento', 'id_type')).upper().strip()
                id_number = str(_get_row_value(row_dict, 'numero_doc', 'numero de documento', 'id_number')).strip()
                nombres = str(_get_row_value(row_dict, 'nombres', 'primer_nombre', 'first_name', 'nombre')).strip()
                apellidos = str(_get_row_value(row_dict, 'apellidos', 'apellido', 'last_name')).strip()
                fecha_nac_raw = _get_row_value(row_dict, 'fecha_nacimiento', 'fecha de nacimiento', 'birth_date', 'nacimiento')
                sexo_raw = str(_get_row_value(row_dict, 'sexo', 'sex', 'genero', 'género')).upper().strip()

                if not id_number:
                    result['errors'].append(f'Fila {row_idx}: número de documento vacío.')
                    continue

                # Normalizar tipo doc
                valid_types = {'CC', 'CE', 'PA', 'TI', 'RC'}
                if id_type_raw not in valid_types:
                    id_type_raw = 'CC'

                # Normalizar sexo
                sexo_map = {'M': 'M', 'MASCULINO': 'M', 'HOMBRE': 'M',
                            'F': 'F', 'FEMENINO': 'F', 'MUJER': 'F',
                            'NB': 'NB', 'NO BINARIO': 'NB', 'X': 'NB'}
                sex = sexo_map.get(sexo_raw, 'M')

                # Normalizar fecha nacimiento
                birth_date = None
                if fecha_nac_raw:
                    if isinstance(fecha_nac_raw, date):
                        birth_date = fecha_nac_raw if isinstance(fecha_nac_raw, date) and not hasattr(fecha_nac_raw, 'hour') else fecha_nac_raw.date()
                    else:
                        from datetime import datetime
                        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d'):
                            try:
                                birth_date = datetime.strptime(str(fecha_nac_raw).strip(), fmt).date()
                                break
                            except ValueError:
                                pass
                        if birth_date is None:
                            result['errors'].append(f'Fila {row_idx}: fecha de nacimiento inválida: {fecha_nac_raw!r}')
                            continue

                if not birth_date:
                    result['errors'].append(f'Fila {row_idx}: fecha de nacimiento requerida.')
                    continue

                full_name = f'{nombres} {apellidos}'.strip()

                # Campos opcionales
                email = str(_get_row_value(row_dict, 'email', 'correo', 'correo_electronico')).strip()
                phone = str(_get_row_value(row_dict, 'telefono', 'teléfono', 'phone')).strip()
                address = str(_get_row_value(row_dict, 'direccion', 'dirección', 'address')).strip()
                has_disability_raw = str(_get_row_value(row_dict, 'discapacidad', 'has_disability')).lower().strip()
                has_disability = has_disability_raw in ('si', 'sí', 'yes', '1', 'true')
                is_head_raw = str(_get_row_value(row_dict, 'cabeza_hogar', 'cabeza de hogar', 'is_head_of_household')).lower().strip()
                is_head = is_head_raw in ('si', 'sí', 'yes', '1', 'true')

                defaults = dict(
                    full_name=full_name or f'{nombres} {apellidos}'.strip() or id_number,
                    first_name=nombres,
                    last_name=apellidos,
                    birth_date=birth_date,
                    sex=sex,
                    has_disability=has_disability,
                    is_head_of_household=is_head,
                    email=email,
                    phone=phone,
                    address=address,
                )
                try:
                    obj, created = Employee.objects.update_or_create(
                        entity=entity,
                        id_type=id_type_raw,
                        id_number=id_number,
                        defaults=defaults,
                    )
                    if created:
                        result['created'] += 1
                    else:
                        result['updated'] += 1
                except Exception as exc:
                    result['errors'].append(f'Fila {row_idx}: {exc}')

    # ---- Hoja: Educación ----
    edu_sheet = None
    for sheet_name in wb.sheetnames:
        norm = _normalize_col(sheet_name)
        if 'educacion' in norm or 'educación' in norm or 'estudio' in norm or 'formacion' in norm:
            edu_sheet = wb[sheet_name]
            break

    if edu_sheet is None:
        result['warnings'].append('Hoja "Educación" no encontrada; se omite.')
    else:
        rows = list(edu_sheet.rows)
        if len(rows) >= 2:
            headers = [str(c.value) if c.value else '' for c in rows[0]]
            for row in rows[1:]:
                row_dict = {headers[i]: (cell.value if cell.value is not None else '') for i, cell in enumerate(row)}
                id_number = str(_get_row_value(row_dict, 'numero_doc', 'id_number', 'documento')).strip()
                if not id_number:
                    continue
                try:
                    emp = Employee.objects.filter(entity=entity, id_number=id_number).first()
                    if emp is None:
                        continue
                    level_raw = str(_get_row_value(row_dict, 'nivel', 'nivel_educativo', 'level')).upper().strip()
                    level_map = {
                        'PRIMARIA': 'PRIMARIA', 'BACHILLERATO': 'BACHILLERATO',
                        'TECNICO': 'TECNICO', 'TÉCNICO': 'TECNICO',
                        'TECNOLOGO': 'TECNOLOGO', 'TECNÓLOGO': 'TECNOLOGO',
                        'PREGRADO': 'PREGRADO', 'PROFESIONAL': 'PREGRADO',
                        'ESPECIALIZACION': 'ESPECIALIZACION', 'ESPECIALIZACIÓN': 'ESPECIALIZACION',
                        'MAESTRIA': 'MAESTRIA', 'MAESTRÍA': 'MAESTRIA',
                        'DOCTORADO': 'DOCTORADO',
                    }
                    level = level_map.get(level_raw, 'PREGRADO')
                    inst = str(_get_row_value(row_dict, 'institucion', 'institución', 'institution')).strip()
                    program = str(_get_row_value(row_dict, 'programa', 'program')).strip()
                    title = str(_get_row_value(row_dict, 'titulo', 'título', 'title')).strip()
                    EmployeeEducation.objects.get_or_create(
                        employee=emp,
                        level=level,
                        program=program or '—',
                        defaults=dict(institution=inst or '—', title=title or '—'),
                    )
                except Exception:
                    pass

    # ---- Hoja: Experiencia ----
    exp_sheet = None
    for sheet_name in wb.sheetnames:
        norm = _normalize_col(sheet_name)
        if 'experiencia' in norm:
            exp_sheet = wb[sheet_name]
            break

    if exp_sheet is None:
        result['warnings'].append('Hoja "Experiencia" no encontrada; se omite.')
    else:
        from datetime import datetime
        rows = list(exp_sheet.rows)
        if len(rows) >= 2:
            headers = [str(c.value) if c.value else '' for c in rows[0]]
            for row in rows[1:]:
                row_dict = {headers[i]: (cell.value if cell.value is not None else '') for i, cell in enumerate(row)}
                id_number = str(_get_row_value(row_dict, 'numero_doc', 'id_number', 'documento')).strip()
                if not id_number:
                    continue
                try:
                    emp = Employee.objects.filter(entity=entity, id_number=id_number).first()
                    if emp is None:
                        continue
                    employer = str(_get_row_value(row_dict, 'empleador', 'employer', 'empresa')).strip()
                    position_name = str(_get_row_value(row_dict, 'cargo', 'position_name', 'puesto')).strip()
                    sector_raw = str(_get_row_value(row_dict, 'sector')).upper().strip()
                    sector_map = {'PUBLICO': 'PUBLICO', 'PÚBLICO': 'PUBLICO',
                                  'PRIVADO': 'PRIVADO', 'MIXTO': 'MIXTO',
                                  'INDEPENDIENTE': 'INDEPENDIENTE'}
                    sector = sector_map.get(sector_raw, 'PUBLICO')
                    is_pub_raw = str(_get_row_value(row_dict, 'es_sector_publico', 'sector_publico', 'is_public_sector')).lower()
                    is_pub = is_pub_raw in ('si', 'sí', 'yes', '1', 'true') or sector == 'PUBLICO'
                    start_raw = _get_row_value(row_dict, 'fecha_inicio', 'start_date', 'inicio')
                    start_date = None
                    if isinstance(start_raw, date):
                        start_date = start_raw if not hasattr(start_raw, 'hour') else start_raw.date()
                    elif start_raw:
                        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
                            try:
                                start_date = datetime.strptime(str(start_raw).strip(), fmt).date()
                                break
                            except ValueError:
                                pass
                    if not start_date:
                        continue
                    EmployeeExperience.objects.get_or_create(
                        employee=emp,
                        employer=employer or '—',
                        start_date=start_date,
                        defaults=dict(
                            position_name=position_name or '—',
                            sector=sector,
                            is_public_sector=is_pub,
                        ),
                    )
                except Exception:
                    pass

    # ---- Hoja: Capacitación ----
    cap_sheet = None
    for sheet_name in wb.sheetnames:
        norm = _normalize_col(sheet_name)
        if 'capacitacion' in norm or 'capacitación' in norm or 'formacion' in norm:
            cap_sheet = wb[sheet_name]
            break

    if cap_sheet is None:
        result['warnings'].append('Hoja "Capacitación" no encontrada; se omite.')
    else:
        rows = list(cap_sheet.rows)
        if len(rows) >= 2:
            headers = [str(c.value) if c.value else '' for c in rows[0]]
            for row in rows[1:]:
                row_dict = {headers[i]: (cell.value if cell.value is not None else '') for i, cell in enumerate(row)}
                id_number = str(_get_row_value(row_dict, 'numero_doc', 'id_number', 'documento')).strip()
                if not id_number:
                    continue
                try:
                    emp = Employee.objects.filter(entity=entity, id_number=id_number).first()
                    if emp is None:
                        continue
                    topic = str(_get_row_value(row_dict, 'tema', 'topic', 'curso')).strip()
                    hours_raw = _get_row_value(row_dict, 'horas', 'hours')
                    try:
                        hours = int(float(str(hours_raw))) if hours_raw else 0
                    except (ValueError, TypeError):
                        hours = 0
                    inst = str(_get_row_value(row_dict, 'institucion', 'institución', 'institution')).strip()
                    from apps.talento.models import EmployeeTraining
                    EmployeeTraining.objects.get_or_create(
                        employee=emp,
                        topic=topic or '—',
                        institution=inst or '—',
                        defaults=dict(hours=hours),
                    )
                except Exception:
                    pass

    return result


def calculate_tenure(employee: 'Employee') -> dict:
    """
    Calcula el tiempo de servicio del empleado.

    Suma:
    - Experiencias con is_public_sector=True.
    - Registros de vinculación activos.

    Retorna: {total_days, days_in_current_entity, total_years}
    """
    today = date.today()
    total_days = 0

    # Experiencias en sector público
    for exp in employee.experience.filter(is_public_sector=True):
        end = exp.end_date or today
        delta = (end - exp.start_date).days
        if delta > 0:
            total_days += delta

    # Vinculaciones activas en la entidad del empleado
    days_in_entity = 0
    for rec in employee.employment_records.filter(is_active=True, entity=employee.entity):
        delta = (today - rec.appointment_date).days
        if delta > 0:
            days_in_entity += delta
            total_days += delta

    return {
        'total_days': total_days,
        'days_in_current_entity': days_in_entity,
        'total_years': round(total_days / 365.25, 2),
    }


def calculate_retirement_eligibility(employee: 'Employee') -> dict:
    """
    Calcula si el empleado es pre-pensionado según la Ley 100/93 ajustada
    por la Ley 797/2003.

    Régimen aplicado (sin RPM especiales):
    - Mujer: 57 años + 1300 semanas cotizadas (≈ 25.01 años).
    - Hombre: 62 años + 1300 semanas cotizadas.

    Pre-pensionado: si le faltan 3 años o menos para cumplir CUALQUIERA de
    los dos requisitos.

    Años cotizados ≈ suma de (end_date or today − start_date).days / 365.25
    para experiencias con is_public_sector=True.

    Retorna: {is_pre_pensioned: bool, years_remaining: float, reason: str}
    """
    today = date.today()

    # Edad actual
    age_years = (today - employee.birth_date).days / 365.25

    # Años cotizados (sector público)
    cotized_days = 0
    for exp in employee.experience.filter(is_public_sector=True):
        end = exp.end_date or today
        delta = (end - exp.start_date).days
        if delta > 0:
            cotized_days += delta
    cotized_years = cotized_days / 365.25

    # Requisitos según sexo
    is_female = employee.sex == 'F'
    age_required = 57.0 if is_female else 62.0
    cotized_required = 25.01  # 1300 semanas / 52

    years_to_age = max(0.0, age_required - age_years)
    years_to_cotized = max(0.0, cotized_required - cotized_years)

    years_remaining = min(years_to_age, years_to_cotized)

    is_pre_pensioned = years_remaining <= 3.0

    if years_remaining == 0:
        reason = 'Cumple requisitos de pensión.'
    elif is_pre_pensioned:
        reason = (
            f'Pre-pensionado: faltan {years_remaining:.1f} años para pensión. '
            f'Edad actual: {age_years:.1f} años (requiere {age_required}), '
            f'semanas cotizadas ≈{cotized_years:.1f} años (requiere {cotized_required}).'
        )
    else:
        reason = (
            f'No pre-pensionado: faltan {years_remaining:.1f} años. '
            f'Edad actual: {age_years:.1f} años, cotizados ≈{cotized_years:.1f} años.'
        )

    return {
        'is_pre_pensioned': is_pre_pensioned,
        'years_remaining': round(years_remaining, 2),
        'reason': reason,
    }
